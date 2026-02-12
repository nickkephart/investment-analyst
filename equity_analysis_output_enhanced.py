#!/usr/bin/env python3
"""
Enhanced Equity Analysis Output Generator
Adds performance metrics and creates structured data table
"""

import csv
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List


class EquityAnalysisOutputGenerator:
    """Generate structured output with performance metrics"""
    
    def __init__(self):
        self.performance_fields = [
            'current_price',
            '52_week_high',
            '52_week_low',
            'pct_from_52w_high',
            'pct_from_52w_low',
            '1_year_return',
            '3_month_return',
            '1_month_return',
            'ytd_return',
            'dividend_yield',
            'market_cap',
            'pe_ratio',
            'volume_avg'
        ]
    
    def generate_csv_output(self, analysis_results: Dict, output_path: str):
        """Generate CSV file with analysis and performance data"""
        
        securities = analysis_results.get('securities', [])
        
        # Define CSV columns
        columns = [
            'Rank',
            'Ticker',
            'Type',
            'Name',
            'Alignment Score',
            'Current Price',
            '52W High',
            '52W Low',
            '% From 52W High',
            '% From 52W Low',
            '1Y Return %',
            '3M Return %',
            '1M Return %',
            'YTD Return %',
            'Div Yield %',
            'Market Cap',
            'P/E Ratio',
            'Avg Volume',
            'Sector',
            'Industry',
            'Revenue Exposure %',
            'Rationale'
        ]
        
        # Write CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            
            for i, sec in enumerate(securities, 1):
                perf = sec.get('performance', {})
                
                row = {
                    'Rank': i,
                    'Ticker': sec.get('ticker', ''),
                    'Type': sec.get('asset_type', 'Stock'),
                    'Name': sec.get('name', ''),
                    'Alignment Score': sec.get('alignment_score', 0),
                    'Current Price': perf.get('current_price', ''),
                    '52W High': perf.get('week_52_high', ''),
                    '52W Low': perf.get('week_52_low', ''),
                    '% From 52W High': perf.get('pct_from_52w_high', ''),
                    '% From 52W Low': perf.get('pct_from_52w_low', ''),
                    '1Y Return %': perf.get('return_1y', ''),
                    '3M Return %': perf.get('return_3m', ''),
                    '1M Return %': perf.get('return_1m', ''),
                    'YTD Return %': perf.get('return_ytd', ''),
                    'Div Yield %': perf.get('dividend_yield', ''),
                    'Market Cap': perf.get('market_cap', ''),
                    'P/E Ratio': perf.get('pe_ratio', ''),
                    'Avg Volume': perf.get('avg_volume', ''),
                    'Sector': sec.get('sector', ''),
                    'Industry': sec.get('industry', ''),
                    'Revenue Exposure %': sec.get('revenue_exposure_pct', ''),
                    'Rationale': sec.get('alignment_rationale', '')
                }
                
                writer.writerow(row)
        
        print(f"âœ“ CSV output saved to: {output_path}")
    
    def generate_excel_output(self, analysis_results: Dict, output_path: str):
        """Generate Excel file with formatted tables"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("âš ï¸  openpyxl not available, skipping Excel output")
            return
        
        securities = analysis_results.get('securities', [])
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Equity Analysis"
        
        # Headers
        headers = [
            'Rank', 'Ticker', 'Type', 'Name', 'Score',
            'Price', '52W High', '52W Low', '% From High', '% From Low',
            '1Y %', '3M %', '1M %', 'YTD %', 'Div %',
            'Market Cap', 'P/E', 'Volume', 'Sector', 'Exposure %', 'Rationale'
        ]
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, sec in enumerate(securities, 2):
            perf = sec.get('performance', {})
            
            # Rank and basic info
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=sec.get('ticker', ''))
            ws.cell(row=row_num, column=3, value=sec.get('asset_type', 'Stock'))
            ws.cell(row=row_num, column=4, value=sec.get('name', ''))
            ws.cell(row=row_num, column=5, value=sec.get('alignment_score', 0))
            
            # Price data
            ws.cell(row=row_num, column=6, value=perf.get('current_price', ''))
            ws.cell(row=row_num, column=7, value=perf.get('week_52_high', ''))
            ws.cell(row=row_num, column=8, value=perf.get('week_52_low', ''))
            ws.cell(row=row_num, column=9, value=perf.get('pct_from_52w_high', ''))
            ws.cell(row=row_num, column=10, value=perf.get('pct_from_52w_low', ''))
            
            # Returns
            ws.cell(row=row_num, column=11, value=perf.get('return_1y', ''))
            ws.cell(row=row_num, column=12, value=perf.get('return_3m', ''))
            ws.cell(row=row_num, column=13, value=perf.get('return_1m', ''))
            ws.cell(row=row_num, column=14, value=perf.get('return_ytd', ''))
            ws.cell(row=row_num, column=15, value=perf.get('dividend_yield', ''))
            
            # Fundamentals
            ws.cell(row=row_num, column=16, value=perf.get('market_cap', ''))
            ws.cell(row=row_num, column=17, value=perf.get('pe_ratio', ''))
            ws.cell(row=row_num, column=18, value=perf.get('avg_volume', ''))
            
            # Classification
            ws.cell(row=row_num, column=19, value=sec.get('sector', ''))
            ws.cell(row=row_num, column=20, value=sec.get('revenue_exposure_pct', ''))
            ws.cell(row=row_num, column=21, value=sec.get('alignment_rationale', ''))
            
            # Color code by score
            score = sec.get('alignment_score', 0)
            if score >= 8.0:
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            elif score >= 6.0:
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
            else:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            
            for col in range(1, 22):
                ws.cell(row=row_num, column=col).fill = fill
        
        # Auto-adjust column widths
        for col_num in range(1, 22):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Make Name and Rationale columns wider
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['U'].width = 50
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(output_path)
        print(f"âœ“ Excel output saved to: {output_path}")


def test_output_generator():
    """Test the output generator with sample data"""
    
    # Sample analysis results
    analysis = {
        'thesis_id': 1,
        'thesis_title': 'AI Infrastructure Buildout',
        'securities': [
            {
                'ticker': 'NVDA',
                'name': 'NVIDIA Corporation',
                'asset_type': 'Stock',
                'alignment_score': 7.7,
                'sector': 'Technology',
                'industry': 'Semiconductors',
                'revenue_exposure_pct': 80,
                'alignment_rationale': 'AI GPU leader, 80%+ data center revenue',
                'performance': {
                    'current_price': 174.19,
                    'week_52_high': 212.19,
                    'week_52_low': 86.62,
                    'pct_from_52w_high': -17.9,
                    'pct_from_52w_low': 101.1,
                    'return_1y': 41.56,
                    'return_3m': -15.2,
                    'return_1m': -8.4,
                    'return_ytd': -5.3,
                    'dividend_yield': 0.02,
                    'market_cap': '4.24T',
                    'pe_ratio': 43.12,
                    'avg_volume': '182M'
                }
            },
            {
                'ticker': 'SOXX',
                'name': 'iShares Semiconductor ETF',
                'asset_type': 'ETF',
                'alignment_score': 8.7,
                'sector': 'Technology',
                'industry': 'ETF - Semiconductors',
                'revenue_exposure_pct': 95,
                'alignment_rationale': 'Pure semiconductor exposure',
                'performance': {
                    'current_price': 225.50,
                    'week_52_high': 275.00,
                    'week_52_low': 180.00,
                    'pct_from_52w_high': -18.0,
                    'pct_from_52w_low': 25.3,
                    'return_1y': 35.2,
                    'return_3m': -10.5,
                    'return_1m': -5.2,
                    'return_ytd': -3.8,
                    'dividend_yield': 0.45,
                    'market_cap': 'N/A',
                    'pe_ratio': 'N/A',
                    'avg_volume': '5.2M'
                }
            }
        ]
    }
    
    # Generate outputs
    generator = EquityAnalysisOutputGenerator()
    
    csv_path = "/mnt/user-data/outputs/equity_analysis_sample.csv"
    generator.generate_csv_output(analysis, csv_path)
    
    excel_path = "/mnt/user-data/outputs/equity_analysis_sample.xlsx"
    generator.generate_excel_output(analysis, excel_path)
    
    print("\nâœ… Sample outputs generated!")
    print(f"   CSV:   {csv_path}")
    print(f"   Excel: {excel_path}")


if __name__ == '__main__':
    test_output_generator()
